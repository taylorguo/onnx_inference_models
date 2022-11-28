'''
Author: Taylor Guo
Date: 2021-05-24 09:45:03
LastEditTime: 2021-05-28 22:27:08
LastEditors: Please set LastEditors
Description: In User Settings Edit
FilePath: /onnx2json/onnx/shape_inference.py
'''

import onnx, json, os
from onnx import shape_inference, numpy_helper
import numpy as np
import openpyxl

def get_onnx_weight(onnx_model):
    # onnx_model = onnx.load(model_path)
    # print(onnx_model.graph.initializer)
    producer = onnx_model.producer_name
    os.system("")
    print("\033[42m"+ "The onnx model comes from {}.".format(producer if producer else "ONNX official"  ) +"\033[0m")
    
    # try:
    #     onnx.checker.check_model(onnx_model)
    # except onnx.checker.ValidationError as e:
    #     print('The model is invalid: %s' % e)
    # else:
    #     print('The model is valid!')

    weight_dict = {}
    weight_list = onnx_model.graph.initializer
    # TODO: pytorch-> onnx : producer= pytorch; onnx official: producer = ""; others to-do
    for n, weight in enumerate(weight_list):
        if "torch" in producer.lower():            
            weight_np = np.frombuffer(weight.raw_data, dtype=np.float32)
        else:  
            weight_np = np.array(weight.float_data)
        weight_dict[weight.name] = weight_np
        # print(n, weight.name, weight_np.shape)
    return weight_dict

######################
# 
def add_value_info_for_constants(model : onnx.ModelProto):
    """
    Currently onnx.shape_inference doesn't use the shape of initializers, so add
    that info explicitly as ValueInfoProtos.
    Mutates the model.
    Args:
        model: The ModelProto to update.
    """
    # All (top-level) constants will have ValueInfos before IRv4 as they are all inputs
    if model.ir_version < 4:
        return

    def add_const_value_infos_to_graph(graph : onnx.GraphProto):
        inputs = {i.name for i in graph.input}
        existing_info = {vi.name: vi for vi in graph.value_info}
        for init in graph.initializer:
            # Check it really is a constant, not an input
            if init.name in inputs:
                continue

            # The details we want to add
            elem_type = init.data_type
            shape = init.dims

            # Get existing or create new value info for this constant
            vi = existing_info.get(init.name)
            if vi is None:
                vi = graph.value_info.add()
                vi.name = init.name

            # Even though it would be weird, we will not overwrite info even if it doesn't match
            tt = vi.type.tensor_type
            if tt.elem_type == onnx.TensorProto.UNDEFINED:
                tt.elem_type = elem_type
            if not tt.HasField("shape"):
                # Ensure we set an empty list if the const is scalar (zero dims)
                tt.shape.dim.extend([])
                for dim in shape:
                    tt.shape.dim.add().dim_value = dim

        # Handle subgraphs
        for node in graph.node:
            for attr in node.attribute:
                # Ref attrs refer to other attrs, so we don't need to do anything
                if attr.ref_attr_name != "":
                    continue

                if attr.type == onnx.AttributeProto.GRAPH:
                    add_const_value_infos_to_graph(attr.g)
                if attr.type == onnx.AttributeProto.GRAPHS:
                    for g in attr.graphs:
                        add_const_value_infos_to_graph(g)


    return add_const_value_infos_to_graph(model.graph)      
###### Shape Inference

def onnx_shape_inference(model_path):
    onnx_model = onnx.load(model_path)
    value_info_list = onnx_model.graph.value_info
    tensor_shape = {}

    try:
        onnx.checker.check_model(onnx_model)
    except onnx.checker.ValidationError as e:
        print('The model is invalid: %s' % e)
        return tensor_shape
    # else:
    #     print('The model is valid!')
    file_prefix = ""
    file_path, file_name = os.path.split(model_path)
    file_prefix, file_ext = os.path.splitext(file_name)
    print("\033[43m"+ "Porcessing Model: {}.".format(file_prefix) +"\033[0m")
    if not len(value_info_list):
        # print("The model contains some shapes, do not apply shape_inference!")
        onnx_model = shape_inference.infer_shapes(onnx_model)
        # onnx.save(onnx_model, model_path+"_inferred.onnx")
        # Check the model and print Y's shape information
        onnx.checker.check_model(onnx_model)
        value_info_list = onnx_model.graph.value_info
        # print('After shape inference, the shape info  is:\n{}'.format(value_info_list)) 
        onnx.save(onnx_model, file_path + "/%s_shapeinference.onnx"%(file_prefix))
        print('Finish shape inference') 

    # Process graph input as Conv input if 2'nd layer is convolution
    graph_input = onnx_model.graph.input[0]

    # get weight shape
    wt_dict = {}
    weights = onnx_model.graph.initializer
    for weight in weights:
        np_wt = numpy_helper.to_array(weight)
        wt_dict[weight.name] = np_wt.shape
    # get nodes input list and output list
    nodes_in_out_dict = {}
    nodes = onnx_model.graph.node
    for node in nodes:
        # {'resnetv22_batchnorm0_fwd': 
        #   {'input': ['data', 'resnetv22_batchnorm0_gamma', 'resnetv22_batchnorm0_beta', 
        #              'resnetv22_batchnorm0...nning_mean', 'resnetv22_batchnorm0...unning_var'], 
        #    'output': ['resnetv22_batchnorm0_fwd']
        #   }
        # }
        node_dict, input_list, output_list = {}, [], []
        node_dict["op_type"] = node.op_type

        for input in node.input:
            input_list.append(input)
        node_dict["input"] = input_list

        for output in node.output:
            output_list.append(output)
        node_dict["output"] = output_list

        nodes_in_out_dict[node.name] = node_dict

    # print(graph_input)
    # if onnx_model.graph.node[1].op_type == "Conv":
    #     shape_list = []
    #     for dv in graph_input.type.tensor_type.shape.dim:
    #         shape_list.append(dv.dim_value)
    #     tensor_shape[graph_input.name] = shape_list
    # shape_list = []
    # for dv in graph_input.type.tensor_type.shape.dim:
    #     shape_list.append(dv.dim_value)
    # tensor_shape[graph_input.name] = shape_list
    # print("**********  value_info_list: ", value_info_list)
    for num, node in enumerate(value_info_list):        
        shape_list = []
        for kv in node.type.tensor_type.shape.dim:
            shape_list.append(kv.dim_value)
        tensor_shape[node.name] = shape_list
    # print(" ########### graph: ", graph_input.name, onnx_model.graph.output[0].name)
    return tensor_shape, get_onnx_weight(onnx_model), file_prefix, nodes_in_out_dict, wt_dict

def read_json_ops_tensors(json_file):
    with open(json_file, "r", encoding="utf-8") as jf:
        json_content = json.load(jf)
    return json_content["Operators"], json_content["Tensors"]

def get_weight_shapes(json_file):
    with open(json_file, "r", encoding="utf-8") as jf:
        json_content = json.load(jf)
    operators = json_content["Operators"]
    tensors = json_content["Tensors"]
    weight_op = {}
    for conv, params in operators.items():
        for name, tensor in params["consumers"].items():
            if (conv in tensor) and ("weight" in tensor):
                weight_op[name] = tensor
    weight_shape = {}
    for wt, pams in tensors.items():
        for wtname, ts in weight_op.items():
            if wt == ts:
                weight_shape[wtname] = pams["shape"]
    # print(weight_shape)
    return weight_shape

def update_json_shapes(ops, tensors, all_tensors):
    input_name, output_name = None, None
    for op, params in ops.items():
        activation_input= params["inputs"][0] 
        activation_output = params["output"][0] 
        for kc, vc in params["consumers"].items():
            if vc == activation_input: input_name = kc
        for kp, vp in params["producers"].items():
            if vp == activation_output: output_name = kp
        # print(input_name, "             ", output_name, "               ", 
        #         activation_input, "             ", activation_output)
        
        for tensor_name, shape_list in all_tensors.items():
            
            if tensor_name in input_name :
                tensors[activation_input]["shape"] = shape_list
            if tensor_name in output_name:
                tensors[activation_output]["shape"] = shape_list
    # print(tensors)

def write_json_shapes(json_file, tensors):
    with open(json_file, "r", encoding="utf-8") as jf:
        json_content = json.load(jf)
        json_content["Tensors"] = tensors

    with open(json_file, "w", encoding="utf-8") as jf:
        jf.write(json.dumps(json_content))


def onnx2json_pypost(model_path, json_file):
    
    all_shapes, onnx_weight = onnx_shape_inference(model_path)
    
    ops, tensors = read_json_ops_tensors(json_file)
    update_json_shapes(ops, tensors, all_shapes)
    write_json_shapes(json_file, tensors)
    # print("Saved tensor list to %s"%(json_file))
    return onnx_weight

######### Get weights with numpy ndarray shape
def get_weights_reshape(model_path, onnx_weight):
    # onnx_weight = get_onnx_weight(onnx_model)
    weight_shape = get_weight_shapes(model_path + ".json")
    ret_weight = {}
    for wtname, wt in onnx_weight.items():
        for wtsnm, shape in weight_shape.items():
            if wtsnm == wtname:
                # print(wtsnm, shape, wt.shape)
                wt_reshape = wt.reshape(shape)
                ret_weight[wtsnm] = wt_reshape
    return ret_weight


def onnx_tool_shape_inference(model_path):
    model = onnx.load_model(model_path)
    producer = model.producer_name
    os.system("")
    print("\033[42m"+ "The onnx model comes from {}.".format(producer if producer else "ONNX official"  ) +"\033[0m")

    file_path, file_name = os.path.split(model_path)
    file_prefix, file_ext = os.path.splitext(file_name)
    
    import onnx_tool
    onnx_tool.model_shape_infer(model, None, saveshapesmodel=file_path + "/%s_shapeinference_.onnx"%(file_prefix))
    print('Finish shape inference') 


# 打印节点信息
def save_node(activation_shape_dict, model_name, nodes_in_out_dict=None, wt_dict=None, excel_path="model_op_list.xlsx"):
    # {
    #  'resnetv22_batchnorm0_fwd': [1, 3, 224, 224], 
    #  'resnetv22_conv0_fwd': [1, 64, 112, 112], 
    #  'resnetv22_batchnorm1_fwd': [1, 64, 112, 112], 
    #  'resnetv22_relu0_fwd': [1, 64, 112, 112], 
    #  'resnetv22_pool0_fwd': [1, 64, 56, 56], 
    #  'resnetv22_stage1_bat...hnorm0_fwd': [1, 64, 56, 56], 
    #  'resnetv22_stage1_activation0': [1, 64, 56, 56], 
    #  'resnetv22_stage1_conv0_fwd': [1, 64, 56, 56], 
    #  'resnetv22_stage1_bat...hnorm1_fwd': [1, 64, 56, 56], 
    #  'resnetv22_stage1_activation1': [1, 64, 56, 56], 
    #  'resnetv22_stage1_conv1_fwd': [1, 64, 56, 56], 
    #  'resnetv22_stage1__plus0': [1, 64, 56, 56], 
    #  'resnetv22_stage1_bat...hnorm2_fwd': [1, 64, 56, 56], 
    #  'resnetv22_stage1_activation2': [1, 64, 56, 56], 
    #  ...
    # }

    # create excel object and sheet object
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet()
    ws.title = model_name

    row0 = ["No.", "op_type", "op_name", "weight_name", "weight_shape", "input_shape", "output_shape"]
    for n, item in enumerate(row0):
        ws.cell(row=1, column=n+1, value=item)

    idx = 1
    for n, (op_name, in_out_dict) in enumerate(nodes_in_out_dict.items()):
        ##### idx, op_type, op_name, weight_name, weight_shape, input_shape, output_shape
        # {'resnetv22_batchnorm0_fwd': {'input': [...], 'output': [...]}}
        # activation_shape_dict = {'resnetv22_batchnorm0_fwd': [1, 3, 224, 224], 
        #                          'resnetv22_conv0_fwd': [1, 64, 112, 112], 
        #                          'resnetv22_batchnorm1_fwd': [1, 64, 112, 112], ... 
        #                          }
        #
        # wt_dict = {'resnetv22_batchnorm0_gamma': (3, ), 'resnetv22_stage1_conv0_weight': (64, 64, 3, 3), ...}
        weight_name, weight_shape, input_shape, output_shape = "", "", "", ""
        op_type = in_out_dict["op_type"]
        in_act, out_act = "", ""

        if op_type == "Conv":
            # activation/input is in position [0], others are weights
            for output_n, out_act in enumerate(in_out_dict["output"]):
                if out_act in activation_shape_dict.keys():
                    output_shape = ",".join(map(str, activation_shape_dict[out_act]))

            for input_n, in_act in enumerate(in_out_dict["input"]):
                if input_n == 0:
                    if in_act in activation_shape_dict.keys():
                        # activation
                        input_shape = ",".join(map(str, activation_shape_dict[in_act]))
                else:
                    SEP_XLS = ", " if input_n > 1 else ""
                    weight_name = in_act
                    if weight_name in wt_dict.keys():
                        weight_shape = ",".join(map(str, wt_dict[weight_name] ))

            # write in Excel, only weight
            ws.cell(row=idx+1, column=1, value = idx)
            ws.cell(row=idx+1, column=2, value = op_type)
            ws.cell(row=idx+1, column=3, value = op_name)
            ws.cell(row=idx+1, column=4, value = weight_name)
            ws.cell(row=idx+1, column=5, value = weight_shape)
            ws.cell(row=idx+1, column=6, value = input_shape)
            ws.cell(row=idx+1, column=7, value = output_shape)
            idx += 1

        elif in_out_dict["op_type"] == "Add":
            # "Add" has no weight
            shape_list = []
            for output_n, out_act in enumerate(in_out_dict["output"]):
                if out_act in activation_shape_dict.keys():
                    output_shape_t = ",".join(map(str, activation_shape_dict[out_act]))
                    shape_list.append(output_shape_t)
            output_shape = " _ ".join(shape_list)
            shape_list = []
            for input_n, in_act in enumerate(in_out_dict["input"]):
                if in_act in activation_shape_dict.keys():
                    # activation
                    input_shape_t = ",".join(map(str, activation_shape_dict[in_act]))
                    shape_list.append(input_shape_t)
            input_shape = " _ ".join(shape_list)
            
            # write in Excel, only weight
            ws.cell(row=idx+1, column=1, value = idx)
            ws.cell(row=idx+1, column=2, value = op_type)
            ws.cell(row=idx+1, column=3, value = op_name)
            ws.cell(row=idx+1, column=4, value = weight_name)
            ws.cell(row=idx+1, column=5, value = weight_shape)
            ws.cell(row=idx+1, column=6, value = input_shape)
            ws.cell(row=idx+1, column=7, value = output_shape)

            idx += 1
        
        else:
            if len(in_out_dict["input"]) > 1:
                wt_name_list, wt_shape_list = [], []
                for num, wt_name in enumerate(in_out_dict["input"]):
                    if num > 0 and wt_name in wt_dict.keys():
                        wt_name_list.append(wt_name)
                        wt_shape_list.append(",".join(map(str, wt_dict[wt_name])))
                # weight_name = in_out_dict["input"][1]
                # weight_shape = ",".join(map(str, wt_dict[weight_name] ))
                weight_name = ", ".join(wt_name_list)
                weight_shape = ", ".join(wt_shape_list)
            if len(in_out_dict["input"]): in_act = in_out_dict["input"][0]
            if len(in_out_dict["output"]): out_act = in_out_dict["output"][0]

            if in_act in activation_shape_dict.keys():
                input_shape = ",".join(map(str, activation_shape_dict[in_act]))
            if out_act in activation_shape_dict.keys():
                output_shape = ",".join(map(str, activation_shape_dict[out_act]))

            # write in Excel, only weight
            ws.cell(row=idx+1, column=1, value = idx)
            ws.cell(row=idx+1, column=2, value = op_type)
            ws.cell(row=idx+1, column=3, value = op_name)
            ws.cell(row=idx+1, column=4, value = weight_name)
            ws.cell(row=idx+1, column=5, value = weight_shape)
            ws.cell(row=idx+1, column=6, value = input_shape)
            ws.cell(row=idx+1, column=7, value = output_shape)

            idx += 1

    wb.save(excel_path)
    print("Saved to excel file : %s"%(excel_path))


if __name__ == "__main__":
    import sys
    # model_path = sys.argv[1]  
    json_file = "build/resnet18-v2-7.onnx.json"
    # json_file = "../conv2_2.onnx.json"

    # onnx2json_pypost(model_path, json_file)
    # print(get_weight_shapes(json_file))
    model_path = "/public/ai_platform/model_coverage/resnet18_shapeinference.onnx"
    model_path = "/public/ai_platform/models/resnet18-v2-7.onnx_inferred.onnx"
    model_path = "D:/vbox/resnet18-v2-7_inferred.onnx"
    # tensor_dict, _, file_prefix, nodes_in_out_dict, wt_dict = onnx_shape_inference(model_path)
    # save_node(tensor_dict, file_prefix, nodes_in_out_dict=nodes_in_out_dict, wt_dict=wt_dict)
    # onnx_tool_shape_inference(model_path)

    model_dir = "/public/ai_platform/model_coverage"
    for root, dirs, files in os.walk(model_dir):
        for file in files:
            file_path = os.path.join(root, file)
            if file_path.endswith(".onnx"):
                # onnx_shape_inference(file_path)
                tensor_dict, _, file_prefix, nodes_in_out_dict, wt_dict = onnx_shape_inference(file_path)
                save_node(tensor_dict, file_prefix, nodes_in_out_dict=nodes_in_out_dict, wt_dict=wt_dict)
    